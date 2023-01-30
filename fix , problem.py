import os
from openpyxl import load_workbook

# Get the file path
file_path = input("Enter the file path (including file name and extension): ")

# Check if the file exists
if not os.path.isfile(file_path):
    print(f"{file_path} does not exist or is not a file.")
    exit()

# Try to load the workbook
try:
    wb = load_workbook(file_path)
except Exception as e:
    print(f"An error occurred while loading the workbook: {e}")
    exit()

# Select the first sheet
ws = wb.active

numeric_list = []

# iterate through the rows in the 'M' column
for row in ws['M']:
    if row.value is not None:
        try:
            numeric_list.append(float(row.value.replace(",", ".")))
        except ValueError:
            numeric_list.append(0)

# Check if the list is not empty
if not numeric_list:
    print("The list of scores is empty.")
    exit()

# calculate the percentage of students who scored under 5/10
under_5 = len([x for x in numeric_list if x < 5]) / len(numeric_list) * 100

# calculate the percentage of students who scored between 5/10 and 7.5/10
between_5_and_7_5 = len([x for x in numeric_list if x >= 5 and x <= 7.5]) / len(numeric_list) * 100

# calculate the percentage of students who scored above 7.5/10
above_7_5 = len([x for x in numeric_list if x > 7.5]) / len(numeric_list) * 100

under_5 = round(under_5, 2)
between_5_and_7_5 = round(between_5_and_7_5, 2)
above_7_5 = round(above_7_5, 2)

# print the results
print("Percentage of students who scored under 5/10:", under_5)
print("Percentage of students who scored between 5/10 and 7.5/10:", between_5_and_7_5)
print("Percentage of students who scored above 7.5/10:", above_7_5)
