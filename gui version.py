from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox, filedialog

def browse_file():
    # open the file browser
    filename = filedialog.askopenfilename(initialdir = "/", title = "Select a file", filetypes = (("Excel files", "*.xlsx"), ("all files", "*.*")))
    if filename:
        # load the workbook
        wb = load_workbook(filename)
        # select the first sheet
        ws = wb.active
        # create a list to store the student scores
        scores = []
        # iterate through the rows in the column specified by the user
        for row in ws[column_letter.get()]:
            if isinstance(row.value, (int, float)):
                scores.append(row.value)
        # calculate the percentage of students who scored under 5/10
        under_5 = len([x for x in scores if x < 5]) / len(scores) * 100
        # calculate the percentage of students who scored between 5/10 and 7.5/10
        between_5_and_7_5 = len([x for x in scores if x >= 5 and x <= 7.5]) / len(scores) * 100
        # calculate the percentage of students who scored above 7.5/10
        above_7_5 = len([x for x in scores if x > 7.5]) / len(scores) * 100
        # round the result to 2 decimal places
        under_5 = round(under_5, 2)
        between_5_and_7_5 = round(between_5_and_7_5, 2)
        above_7_5 = round(above_7_5, 2)
        return under_5, between_5_and_7_5, above_7_5
    else:
        return


# function to display the results in a GUI window
def show_results():
    # call the browse_file() function to get the results
    under_5, between_5_and_7_5, above_7_5 = browse_file()
    if under_5 and between_5_and_7_5 and above_7_5:
        # display the results in the same window using the messagebox.showinfo() method
        messagebox.showinfo("Results", f"Percentage of students who scored under 5/10: {under_5}%\nPercentage of students who scored between 5/10 and 7.5/10: {between_5_and_7_5}%\nPercentage of students who scored above 7.5/10: {above_7_5}%")
    else:
        if not column_letter.get():
            messagebox.showinfo("Results", "Please enter a column letter")
        
# create the GUI window
root = tk.Tk()
root.title("Student Scores")

# create a label for the text field
column_letter_label = tk.Label(root, text="Enter column letter:")
# Create a variable to store the column letter
column_letter = tk.StringVar()

# Create a label to prompt the user to enter the column letter
label = tk.Label(root, text="Enter the column letter:")
label.pack()

# Create a text field for the user to enter the column letter
text_field = tk.Entry(root, textvariable=column_letter)
text_field.pack()

# Create a button to display the results
btn = tk.Button(root, text="Show Results", command=show_results)
btn.pack()

# Run the GUI event loop
root.mainloop()
